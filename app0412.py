import streamlit as st
import pandas as pd
from io import BytesIO
from core import process_all

st.set_page_config(page_title="PRER数据分析系统", layout="wide")

# =========================
# 初始化状态
# =========================
if "eff_results" not in st.session_state:
    st.session_state.eff_results = None

if "saf_results" not in st.session_state:
    st.session_state.saf_results = None

if "has_result" not in st.session_state:
    st.session_state.has_result = False


# =========================
# 页面标题
# =========================
st.title("📊 PRER 文献数据自动分析系统")

# =========================
# 左下角版权
# =========================
st.markdown(
    """
    <style>
    .footer {
        position: fixed;
        left: 10px;
        bottom: 10px;
        color: gray;
        font-size: 12px;
    }
    </style>
    <div class="footer">
        CER中心——数据分析工具
    </div>
    """,
    unsafe_allow_html=True
)

# =========================
# 左右布局
# =========================
left, right = st.columns([1, 2])


# =========================
# Excel转Bytes
# =========================
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def to_excel_bytes(results_dict):

    buffer = BytesIO()

    # 1️⃣ 写入Excel
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, df in results_dict.items():
            df.to_excel(writer, sheet_name=str(name)[:31], index=False)

    buffer.seek(0)

    # 2️⃣ 读取并处理（关键）
    wb = load_workbook(buffer)

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # ===== 纵向合并 =====
        for col in range(1, ws.max_column + 1):
            start = None

            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value

                if val == "合并计算":
                    if start is None:
                        start = row
                else:
                    if start is not None:
                        ws.merge_cells(start_row=start, start_column=col,
                                       end_row=row-1, end_column=col)
                        start = None

            if start is not None:
                ws.merge_cells(start_row=start, start_column=col,
                               end_row=ws.max_row, end_column=col)

        # ===== 横向合并 =====
        for row in range(2, ws.max_row + 1):

            start = None

            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value

                if val == "合并计算":
                    if start is None:
                        start = col
                else:
                    if start is not None:
                        ws.merge_cells(start_row=row, start_column=start,
                                       end_row=row, end_column=col-1)
                        start = None

            if start is not None:
                ws.merge_cells(start_row=row, start_column=start,
                               end_row=row, end_column=ws.max_column)

        # ===== 居中 =====
        for row in ws.iter_rows():
            for c in row:
                c.alignment = Alignment(horizontal="center", vertical="center")

        # ===== 自动列宽 =====
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter  # 列字母

    for cell in col:
        try:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
        except:
            pass

    # 👉 核心：宽度计算（中文适配）
    adjusted_width = max_length + 2

    # 可选：限制最大宽度（防止超长）
    if adjusted_width > 50:
        adjusted_width = 50

    ws.column_dimensions[col_letter].width = adjusted_width

    # 3️⃣ 写回buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


# =========================
# 左侧：输入区
# =========================
with left:

    st.header("📂 数据上传")

    eff_file = st.file_uploader("有效性数据", type=["xlsx"])
    saf_file = st.file_uploader("安全性数据", type=["xlsx"])

    st.divider()

    # =========================
    # 按钮区
    # =========================
    if st.button("🚀 开始分析", use_container_width=True):

        if eff_file is None or saf_file is None:
            st.error("请上传两个Excel文件")
        else:
            try:
                with st.spinner("正在分析中..."):

                    eff_results, saf_results = process_all(
                        eff_file,
                        saf_file
                    )

                    st.session_state.eff_results = eff_results
                    st.session_state.saf_results = saf_results
                    st.session_state.has_result = True

                st.success("✅ 分析完成")

            except Exception as e:
                st.error("❌ 运行失败")
                st.exception(e)

    # =========================
    # 重置按钮（新增）
    # =========================
    if st.session_state.has_result:
        if st.button("🔄 重新分析", use_container_width=True):
            st.session_state.eff_results = None
            st.session_state.saf_results = None
            st.session_state.has_result = False
            st.rerun()


# =========================
# 右侧：结果区
# =========================
with right:

    st.header("📊 分析结果")

    if not st.session_state.has_result:
        st.info("请上传数据并点击“开始分析”")
    else:
        st.success("📌 当前结果已生成")

        eff_results = st.session_state.eff_results
        saf_results = st.session_state.saf_results

        # =========================
        # 下载区
        # =========================
        st.subheader("📥 下载结果")

        eff_excel = to_excel_bytes(eff_results)
        saf_excel = to_excel_bytes(saf_results)

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("**有效性结果** ✅")
            st.download_button(
                "⬇️ 下载",
                data=eff_excel,
                file_name="有效性_结果.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with col2:
            st.markdown("**安全性结果** ✅")
            st.download_button(
                "⬇️ 下载",
                data=saf_excel,
                file_name="安全性_结果.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        st.divider()

        # =========================
        # 预览区
        # =========================
        st.subheader("📊 数据预览")

        tab1, tab2 = st.tabs(["有效性", "安全性"])

        with tab1:
            st.dataframe(list(eff_results.values())[0], use_container_width=True)

        with tab2:
            st.dataframe(list(saf_results.values())[0], use_container_width=True)
