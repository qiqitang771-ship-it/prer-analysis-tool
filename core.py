import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os


# =========================
# 数值清洗
# =========================
def to_numeric_safe(df, cols):
    for c in cols:
        if c in df.columns:
            df[c] = (
                df[c].astype(str)
                .str.replace(",", "", regex=False)
                .str.replace("—", "", regex=False)
                .str.replace("NA", "", regex=False)
                .str.replace("nan", "", regex=False)
                .str.strip()
            )
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


# =========================
# 字段补齐
# =========================
def ensure_columns(df, cols, fill="NA"):
    for c in cols:
        if c not in df.columns:
            df[c] = fill
    return df


# =========================
# 数据读取
# =========================
def load_data(eff_path, saf_path):

    eff_map = {
        "文献编号": "doc_id",
        "组别": "group",
        "有效性指标": "outcome",
        "访视点": "timepoint",
        "器械": "device",
        "数据类型": "type",
        "样本量": "n",
        "均值": "mean",
        "标准差": "sd",
        "组内P值": "p_in",
        "组间P值": "p_between",
        "发生例数": "event"
    }

    saf_map = {
        "文献编号": "doc_id",
        "组别": "group",
        "器械": "device",
        "安全性指标分类": "category",
        "安全性指标": "outcome",
        "数据类型": "type",
        "样本量": "n",
        "发生例数": "event"
    }

    df_eff = pd.read_excel(eff_path)
    df_saf = pd.read_excel(saf_path)

    df_eff.columns = df_eff.columns.str.strip()
    df_saf.columns = df_saf.columns.str.strip()

    df_eff = df_eff.rename(columns=eff_map)
    df_saf = df_saf.rename(columns=saf_map)

    df_eff = ensure_columns(df_eff, ["group", "device", "timepoint"])
    df_saf = ensure_columns(df_saf, ["group", "device", "category"])

    df_eff = to_numeric_safe(df_eff, ["n", "mean", "sd", "event"])
    df_saf = to_numeric_safe(df_saf, ["n", "event"])

    return df_eff, df_saf


# =========================
# 合并计算
# =========================
def pooled_continuous(df):
    df = df.dropna(subset=["n", "mean", "sd"])
    df = df[df["n"] > 0]

    n_total = df["n"].sum()
    if n_total == 0:
        return 0, np.nan, np.nan

    mean_total = np.sum(df["mean"] * df["n"]) / n_total

    ss = np.sum(
        (df["n"] - 1)*(df["sd"]**2) +
        df["n"]*(df["mean"] - mean_total)**2
    )

    sd_total = np.sqrt(ss/(n_total-1)) if n_total > 1 else 0

    return n_total, mean_total, sd_total


def pooled_binary(df):
    df = df.dropna(subset=["n", "event"])
    df = df[df["n"] > 0]

    total_n = df["n"].sum()
    total_event = df["event"].sum()

    rate = total_event / total_n if total_n > 0 else 0

    return total_n, total_event, rate


# =========================
# 有效性
# =========================
def build_eff_table(sub):

    sub = ensure_columns(sub, ["doc_id", "group", "device", "timepoint", "n", "mean", "sd", "p_in", "p_between"])

    if sub["type"].iloc[0] == "连续型":

        table = sub[[
            "doc_id", "group", "device", "timepoint",
            "n", "mean", "sd", "p_in", "p_between"
        ]].copy()

        grouped = sub.groupby(["outcome", "device", "timepoint"], dropna=False)

        merge_rows = []

        for (_, _, tp), g in grouped:

            if g["doc_id"].nunique() >= 2:

                n, mean, sd = pooled_continuous(g)

                merge_rows.append({
                    "doc_id": "合并计算",
                    "group": "合并计算",
                    "device": "合并计算",
                    "timepoint": tp,
                    "n": n,
                    "mean": mean,
                    "sd": sd,
                    "p_in": "/",
                    "p_between": "/"
                })

        table = pd.concat([table, pd.DataFrame(merge_rows)], ignore_index=True)

        table = table.rename(columns={
            "doc_id": "文献编号",
            "group": "组别",
            "device": "器械",
            "timepoint": "访视点",
            "n": "样本量",
            "mean": "均值",
            "sd": "标准差",
            "p_in": "组内P值",
            "p_between": "组间P值"
        })

    else:

        sub = ensure_columns(sub, ["doc_id", "group", "device", "n", "event"])

        table = sub[["doc_id", "group", "device", "n", "event"]].copy()
        table["rate"] = table["event"] / table["n"]

        n, event, rate = pooled_binary(sub)

        merge_row = {
            "doc_id": "合并计算",
            "group": "合并计算",
            "device": "合并计算",
            "n": n,
            "event": event,
            "rate": rate
        }

        table = pd.concat([table, pd.DataFrame([merge_row])], ignore_index=True)

        table = table.rename(columns={
            "doc_id": "文献编号",
            "group": "组别",
            "device": "器械",
            "n": "样本量",
            "event": "发生例数",
            "rate": "发生率"
        })

    return table


# =========================
# 安全性
# =========================
def build_safety_table(df, category):

    sub = df[df["category"] == category].copy()
    sub = ensure_columns(sub, ["doc_id", "group", "device", "outcome", "n", "event"])

    sub = sub.dropna(subset=["n", "event"])
    sub = sub[sub["n"] > 0]

    total_n = sub.drop_duplicates(subset=["doc_id", "group", "device"])["n"].sum()

    table = pd.DataFrame({
        "文献编号": sub["doc_id"],
        "组别": sub["group"],
        "器械": sub["device"],
        "安全性指标": sub["outcome"],
        "样本量": sub["n"],
        "发生例数": sub["event"],
        "发生率": sub["event"] / sub["n"]
    })

    merge_rows = []

    for outcome, g in sub.groupby("outcome", dropna=False):

        event_sum = g["event"].sum()

        merge_rows.append({
            "文献编号": "合并计算",
            "组别": "合并计算",
            "器械": "合并计算",
            "安全性指标": f"{outcome}",
            "样本量": total_n,
            "发生例数": event_sum,
            "发生率": event_sum / total_n if total_n > 0 else 0
        })

    total_event = sub["event"].sum()

    merge_rows.append({
        "文献编号": "合并计算",
        "组别": "合并计算",
        "器械": "合并计算",
        "安全性指标": f"总{category}",
        "样本量": total_n,
        "发生例数": total_event,
        "发生率": total_event / total_n if total_n > 0 else 0
    })

    table = pd.concat([table, pd.DataFrame(merge_rows)], ignore_index=True)

    return table


# =========================
# ⭐ Excel输出（升级版：全方向合并“合并计算”）
# =========================
def export_excel(results, path):

    if os.path.exists(path):
        try:
            os.remove(path)
        except:
            print("请关闭Excel文件")
            return

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, table in results.items():
            table.to_excel(writer, sheet_name=str(name)[:31], index=False)

    wb = load_workbook(path)

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # =========================
        # ⭐ 核心：横向 + 纵向全部合并
        # =========================

        # ---- 纵向合并 ----
        for col in range(1, ws.max_column + 1):
            start = None

            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value

                if val == "合并计算":
                    if start is None:
                        start = row
                else:
                    if start is not None:
                        ws.merge_cells(start_row=start, start_column=col,
                                       end_row=row-1, end_column=col)
                        start = None

            if start is not None:
                ws.merge_cells(start_row=start, start_column=col,
                               end_row=ws.max_row, end_column=col)

        # ---- 横向合并 ----
        for row in range(2, ws.max_row + 1):

            start = None

            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value

                if val == "合并计算":
                    if start is None:
                        start = col
                else:
                    if start is not None:
                        ws.merge_cells(start_row=row, start_column=start,
                                       end_row=row, end_column=col-1)
                        start = None

            if start is not None:
                ws.merge_cells(start_row=row, start_column=start,
                               end_row=row, end_column=ws.max_column)

        # ---- 居中 ----
        for row in ws.iter_rows():
            for c in row:
                c.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(path)
    print(f"✔ 输出完成：{path}")


# =========================
# web调用
# =========================
def run(eff_path, saf_path, out_eff, out_saf):

    df_eff, df_saf = load_data(eff_path, saf_path)

    eff_results = {}
    for outcome in df_eff["outcome"].dropna().unique():
        eff_results[outcome] = build_eff_table(df_eff[df_eff["outcome"] == outcome])

    saf_results = {}
    for category in df_saf["category"].dropna().unique():
        saf_results[category] = build_safety_table(df_saf, category)

    export_excel(eff_results, out_eff)
    export_excel(saf_results, out_saf)
